from openpyxl import load_workbook
import datetime
from datetime import date
import os

#enviornment variable that contains the file path to the excel sheet which has the data on indivudals birthdays
BDAYDATAPATH = os.getenv('BDAYDATAPATH')

#opens excel file
Data = load_workbook(BDAYDATAPATH)
Sheet = Data.active

today = datetime.date.today()
cell = Sheet.cell(row = 2 , column = 2)
datevar = cell.value

#File will be sorted in order based on mmdd format and years will not be a factor in sorting
#Function will utilizie recursive binary search to search for if it is someones birthday
#Function will return the row where the birthday persons information starts, or 0 if its no ones birthday
def check(today, first, last):
    col = 2
    cell = Sheet.cell(row = first, column=col)
    firstid = cell.value
    firstdate = datetime.date(today.year, firstid.month, firstid.day)
    cell = Sheet.cell(row = last, column=col)
    lastid = cell.value
    lastdate = datetime.date(today.year, lastid.month, lastid.day)
    mid = int((first + last)/2)
    cell = Sheet.cell(row = mid, column=col)
    mid_id = cell.value
    mid_date = datetime.date(today.year, mid_id.month, mid_id.day)
    if(firstdate > lastdate):
        return 0
    elif(firstdate==today):
        return first
    elif(lastdate==today):
        return last
    elif(mid_date == today):
        return mid
    elif(today<mid_date):
        return check(today, first, mid-1)
    elif(today>mid_date):
        return check(today, mid+1, last)

