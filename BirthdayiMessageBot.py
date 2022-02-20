import datetime
import time
from openpyxl import load_workbook
from datetime import date
import send_iMessage
import BinarySearchDates
import os

#environment variable that contains the file path to the excel sheet which has the data on individuals birthdays
BDAYDATAPATH = os.getenv('BDAYDATAPATH') 

Data = load_workbook(BDAYDATAPATH) 
Sheet = Data.active

#creates an empty method to be used later
class Empty:
    pass

#function that uses excel information to create a person object
def makePerson(person, row_num):
    col = 1

    cell = Sheet.cell(row = row_num, column=col)
    col = col + 1
    person.name = cell.value

    cell = Sheet.cell(row = row_num, column=col)
    col = col + 1
    person.birthday = cell.value

    cell = Sheet.cell(row = row_num, column=col)
    col = col + 1
    person.phone = cell.value

    cell = Sheet.cell(row = row_num, column=col)
    col = col + 1
    person.imsg = cell.value
    
today = datetime.date.today()

#row where birthday person data exists (if row is 0 it is nobodys birthday)
row = BinarySearchDates.check(today, 2, Sheet.max_row)

#creates an empty class to be filled later; represents my info
me = Empty()
cell = Sheet.cell(row = 2, column=6)
me.phone = cell.value

#creates an empty class to be filled later; represents Birthday Persons info
person = Empty()

#if its someones birthday, fill the empty person with their information
#check if the person has imessage; if they do send the birthday message otherwise send reminder to myself
if row !=0:
    makePerson(person, row)
    phone = person.phones
    name = person.name 
    if(person.imsg == True):
        send_iMessage.sendBirthdayMessage(phone, name)
        time.sleep(3)
        send_iMessage.Reminder(me.phone, name)
    else:
        send_iMessage.Reminder(me.phone, name)
else:
    send_iMessage.nothingSpecial(me.phone)
        
        



